import face_recognition
import cv2
from openpyxl import Workbook
import datetime
import os



# Get a reference to webcam #0 (the default one)
video_capture = cv2.VideoCapture(0)
    
    
# Create a woorksheet
book=Workbook()
sheet=book.active
thisdir = os.getcwd()
# known_face_encodings = []
# known_face_names_txt = []
# known_face_names = []
#
# for r,d,f in os.walk(thisdir):
#     counter = 1
#     for file in f:
#         if ".jpeg" in file:
#             image_borhan = face_recognition.load_image_file(os.getcwd() + "/image/"+file)
#             image_borhan_face_encoding = face_recognition.face_encodings(image_borhan)[0]
#             known_face_encodings.append(image_borhan_face_encoding)
#             known_face_names_txt.append( os.path.splitext(file)[0])
#             known_face_names.append(str(counter))
#             print(counter,os.path.splitext(file)[0])
#             counter+=1
# # Load images.
#
image_borhan = face_recognition.load_image_file(os.getcwd()+"/image/borhan.jpeg")
image_borhan_face_encoding = face_recognition.face_encodings(image_borhan)[0]

image_dada = face_recognition.load_image_file(os.getcwd()+"/image/dada.jpeg")
image_dada_face_encoding = face_recognition.face_encodings(image_dada)[0]

image_jason = face_recognition.load_image_file(os.getcwd()+"/image/jason.jpeg")
image_jason_face_encoding = face_recognition.face_encodings(image_jason)[0]

image_soumith = face_recognition.load_image_file(os.getcwd()+"/image/soumith.jpeg")
image_soumith_face_encoding = face_recognition.face_encodings(image_soumith)[0]

image_forhad = face_recognition.load_image_file(os.getcwd()+"/image/forhad.jpeg")
image_forhad_face_encoding = face_recognition.face_encodings(image_forhad)[0]


# Create arrays of known face encodings and their names
known_face_encodings = [

        image_borhan_face_encoding,
        image_dada_face_encoding,
        image_jason_face_encoding,
        image_soumith_face_encoding,
        image_forhad_face_encoding

    ]
known_face_names = [

        "1",
        "2",
        "3",
        "4",
        "5"
    ]

# Initialize some variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True
    
# Load present date and time
now= datetime.datetime.now()
today=now.day
month=now.month
    
   
while True:
 # Grab a single frame of video
    ret, frame = video_capture.read()
    
    # Resize frame of video to 1/4 size for faster face recognition processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    
    # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
    rgb_small_frame = small_frame[:, :, ::-1]
    
    # Only process every other frame of video to save time
    if process_this_frame:
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
    
    face_names = []
    for face_encoding in face_encodings:
        # See if the face is a match for the known face(s)
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"
    
        # If a match was found in known_face_encodings, just use the first one.
        if True in matches:
            first_match_index = matches.index(True)
            name = known_face_names[first_match_index]
            # Assign attendance
            if int(name) in range(1, 61):
                #sheet.cell(row=int(name), column=int(today)).value = "Present"
                sheet.cell(row=int(name), column=int(1)).value = "Present"
            else:
                pass
    
        face_names.append(name)
        #print("   ", name)
    process_this_frame = not process_this_frame
    
    top=0
    right=0
    bottom=0
    left=0
    # Display the results
    original_name = "A"
    for (top, right, bottom, left), name in zip(face_locations, face_names):
           # Scale back up face locations since the frame we detected in was scaled to 1/4 size
           top *= 4
           right *= 4
           bottom *= 4
           left *= 4
           #print(name)
           original_name = name #known_face_names_txt[int(name)]
    
    # Draw a box around the face
    cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

           # Draw a label with a name below the face
    cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
    font = cv2.FONT_HERSHEY_DUPLEX
    cv2.putText(frame, original_name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
    
    # Display the resulting image
    cv2.imshow('Video', frame)
        
    # Save Woorksheet as present month
    book.save(str(month)+'.xlsx')
    
    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break
    
# Release handle to the webcam
video_capture.release()
cv2.destroyAllWindows()
    
   
